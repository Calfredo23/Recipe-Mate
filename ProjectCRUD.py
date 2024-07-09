from tkinter import * 
import random
from tkinter import ttk,Tk,Label,font,messagebox,filedialog
import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl.drawing.image import Image as xlImage
from PIL import Image, ImageTk

#To run this program, you must install these packages:
                #pip install tkinter
                #pip install openpyxl
                #pip install openpyxl_image_loader
                #pip install pillow


loading_interface=tk.Tk()
loading_interface.overrideredirect(1)
loading_interface.attributes('-toolwindow',True)
loading_interface.resizable(FALSE, FALSE)

excel_con=Workbook()
excel_con=load_workbook('RecipeMateDB.xlsx')
excel_activate=excel_con.active
sheet1 = excel_con['Sheet1']

#window configure to center
window_width=350
window_height=350
screen_width=loading_interface.winfo_screenwidth()
screen_height=loading_interface.winfo_screenheight()
x_coordinate=screen_width//2 - window_width//2
y_coordinate=screen_height//2 - window_height//2

#window position
loading_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

def fade_out():
    loading_interface.after(3000, loading_to_login)

def clear_entries_recipe():
    created_name=NameofDish_ent.delete(0,END)
    created_method=mmcombo.set("")
    created_category=cofcombo.set("")
    created_ingredients=ingtxt.delete(1.0,END)
    created_procedure=proctxt.delete(1.0,END)
    
def clear_entries_main():
    main_search_ent.delete(0,END)

#===================================================================================================================================

def loading_to_login():
    loading_interface.destroy()
    login_interface_show()

#===================================================================================================================================

def login_interface_show():
    global login_interface
    login_interface = tk.Tk()
    login_interface.geometry('1200x700')     
    login_interface.overrideredirect(1)
    login_interface.resizable(False, False)           

    global close_window_login
    def close_window_login():
        login_interface.destroy()

    def move_window(event):
        login_interface.geometry(f'+{event.x_root}+{event.y_root}')


    def drag(event):
        login_interface_x = login_interface.winfo_pointerx() - login_interface._offsetx
        login_interface_y = login_interface.winfo_pointery() - login_interface._offsety
        login_interface.geometry(f'+{login_interface_x}+{login_interface_y}')

    def start_drag(event):
        login_interface._offsetx = event.x
        login_interface._offsety = event.y


    # login interface title bar
    login_interface_title_bar=Frame(login_interface, bg='#242031')
    login_interface_title_bar.pack(fill=X)

    login_interface.bind('<Button-1>',start_drag)
    login_interface.bind('<B1-Motion>',drag)

    # login interface title bar icon
    login_interface_title_label=Label(login_interface_title_bar,text='RecipeMate (Beta) 0.0.1',fg='white',bg='#242031')
    login_interface_icon=PhotoImage(file=r'icon.png')
    login_interface_icon_pic=Label(login_interface_title_bar,image=login_interface_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
    login_interface_title_label.pack(side=LEFT, padx=10)
    
    # login interface close button
    login_close_button=Button(login_interface_title_bar,text='✕',command=lambda:close_window_login(),bg='#242031',fg='white')
    login_close_button.pack(side=RIGHT,padx=0,ipadx=7)
    
    #window configure to center
    window_width = 1200
    window_height = 700
    screen_width = login_interface.winfo_screenwidth()
    screen_height = login_interface.winfo_screenheight()
    x_coordinate = screen_width // 2 - window_width // 2
    y_coordinate = screen_height // 2 - window_height // 2

    #window position
    login_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')
    
    #login interface image background
    login_interface_image=Image.open(r'login_interfacebg.png')
    photo_login_interface_image=ImageTk.PhotoImage(login_interface_image)
    labellogin=Label(login_interface,image=photo_login_interface_image)
    labellogin.pack()
    
    login_saying_lbl=Label(login_interface,text='Login to continue',fg='white',bg='#242031',font=('Helvetica',20)).place(x=190,y=250)
    
    #Button with image
    createacc_photo = PhotoImage(file = r'create_acc.png')
    createacc_but=Button(login_interface,image=createacc_photo,borderwidth=0,highlightthickness=0,command=lambda:createacc_interface_show())
    createacc_but.place(x=56,y=550)
    createacc_but.config(activebackground='#242031')
    login_photo = PhotoImage(file = r'login.png')
    login_but=Button(login_interface,image=login_photo,borderwidth=0,highlightthickness=0,command=lambda:login_process())
    login_but.place(x=325,y=550)
    login_but.config(activebackground='#242031')


    def show_password():
        password = password_ent.get()
        if checkif.get()==1:
            password_ent.config(show='')
            password_ent.delete(0, END)
            password_ent.insert(0, password)
        elif checkif.get()==0:
            password_ent.config(show='*')
            
    #label design
    global username_ent
    global password_ent
    username_lbl=Label(login_interface,text='Username',fg='white',bg='#242031',font=('Arial',20)).place(x=80,y=327)
    username_ent=Entry(login_interface,fg='white',bg='#242031',font=('Arial',20),highlightthickness=0,width=31,insertbackground='white')
    username_ent.place(x=80,y=357)
    username_ent.configure(relief='flat')
    password_lbl=Label(login_interface,text='Password',fg='white',bg='#242031',font=('Arial',20)).place(x=80,y=417)
    password_ent=Entry(login_interface,show='*',fg='white',bg='#242031',font=('Arial',20),highlightthickness=0,width=31,insertbackground='white')
    password_ent.place(x=80,y=447)
    password_ent.configure(relief='flat')
    checkif=IntVar()
    show_pass = Checkbutton(login_interface, text='Show Password', variable=checkif,bg='#242031',fg='white',command=lambda:show_password())
    show_pass.place(x=80,y=497)

    login_interface.mainloop()

def login_process():
    global current_username
    current_username=username_ent.get()
    current_password=password_ent.get()
    if username_ent.get() == '' and password_ent.get() != '':
        messagebox.showwarning('Missing','Please Put Your Username')
    if password_ent.get() == '' and username_ent.get() != '':
        messagebox.showwarning('Missing','Please Put Your Password')
    elif password_ent.get() == '' and username_ent.get() == '':
        messagebox.showerror('Missing','Please Check Your Username and Password')
    elif password_ent.get() != '' and username_ent.get() != '':   
        found_account=False
        for each_cell in range(2,sheet1.max_row +1):
            if excel_activate['A'+str(each_cell)].value == current_username and excel_activate['B'+str(each_cell)].value == current_password:
                found_account=True
        if found_account==True:
            messagebox.showinfo('Success','Logged in Successfully')
            main_interface_show()
        else:
            messagebox.showerror('Error','Invalid Username or Password')
                
#===================================================================================================================================

def createacc_interface_show():
    login_interface.withdraw()
    global createacc_interface
    createacc_interface = tk.Toplevel(login_interface)                          
    createacc_interface.geometry('1200x700')     
    createacc_interface.overrideredirect(1)
    createacc_interface.resizable(False, False)

    global close_window_createacc
    def close_window_createacc():
        createacc_interface.destroy()
        login_interface.deiconify()
        
    def move_window(event):
        createacc_interface.geometry(f'+{event.x_root}+{event.y_root}')

    def drag(event):
        createacc_interface_x = createacc_interface.winfo_pointerx() - createacc_interface._offsetx
        createacc_interface_y = createacc_interface.winfo_pointery() - createacc_interface._offsety
        createacc_interface.geometry(f'+{createacc_interface_x}+{createacc_interface_y}')

    def start_drag(event):
        createacc_interface._offsetx = event.x
        createacc_interface._offsety = event.y
        
    # create acc title bar
    createacc_title_bar = Frame(createacc_interface, bg='#242031')
    createacc_title_bar.pack(fill=X)

    createacc_interface.bind('<Button-1>', start_drag)
    createacc_interface.bind('<B1-Motion>', drag)
    
    # create acc interface title bar icon
    createacc_interface_title_label = Label(createacc_title_bar, text='RecipeMate (Beta) 0.0.1', fg='white', bg='#242031')
    createacc_interface_icon=PhotoImage(file=r'icon.png')
    createacc_interface_icon_pic=Label(createacc_title_bar,image=createacc_interface_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
    createacc_interface_title_label.pack(side=LEFT, padx=10)

    # create acc interface close button
    createacc_close_button = Button(createacc_title_bar, text='✕', command=lambda:close_window_createacc(),bg='#242031',fg='white')
    createacc_close_button.pack(side=RIGHT, padx=0, ipadx=7)

    #window configure to center
    window_width = 1200
    window_height = 700
    screen_width = createacc_interface.winfo_screenwidth()
    screen_height = createacc_interface.winfo_screenheight()
    x_coordinate = screen_width // 2 - window_width // 2
    y_coordinate = screen_height // 2 - window_height // 2

    #window position
    createacc_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

    #create acc interface image background
    createacc_interface_image = Image.open(r'login_interfacebg.png')
    photo_createacc_interface_image = ImageTk.PhotoImage(createacc_interface_image)
    labelcreateacc = Label(createacc_interface, image=photo_createacc_interface_image)
    labelcreateacc.pack()

    global create_username_ent
    global create_password_ent
    create_username_lbl=Label(createacc_interface,text='Create Username',fg='white',bg='#242031',font=('Arial',20)).place(x=80,y=327)
    create_username_ent=Entry(createacc_interface,fg='white',bg='#242031',font=('Arial',20),highlightthickness=0,width=31,insertbackground='white')
    create_username_ent.place(x=80,y=357)
    create_username_ent.configure(relief='flat')
    create_password_lbl=Label(createacc_interface,text='Create Password',fg='white',bg='#242031',font=('Arial',20)).place(x=80,y=412)
    create_password_ent=Entry(createacc_interface,fg='white',bg='#242031',font=('Arial',20),highlightthickness=0,width=31,insertbackground='white')
    create_password_ent.place(x=80,y=447)
    create_password_ent.configure(relief='flat')

    createacc_photo = PhotoImage(file = r'create_acc.png')
    createacc_but=Button(createacc_interface,image=createacc_photo,borderwidth=0, highlightthickness=0,command=lambda:create_acc_process())
    createacc_but.place(x=325,y=520)
    createacc_but.config(activebackground='#242031')
    backtologin_photo = PhotoImage(file = r'back_to_login.png')
    backtologin_but=Button(createacc_interface,image=backtologin_photo,borderwidth=0, highlightthickness=0,command=lambda:close_window_createacc())
    backtologin_but.place(x=56,y=520)
    backtologin_but.config(activebackground='#242031')

    caution_text_font=font.Font(family='Helvetica',size=15,slant='italic')
    createacc_saying_lbl=Label(createacc_interface,text='Creating your own Account',fg='white',bg='#242031',font=('Helvetica',20)).place(x=134,y=250)
    caution_lbl=Label(createacc_interface,text='5 Characters Limit',fg='white',bg='#242031',font=caution_text_font).place(x=370,y=320)
    
    

    
    createacc_interface.mainloop()

def create_acc_process():
    login_interface.withdraw()
    global created_username
    global created_password
    created_username=create_username_ent.get()
    created_password=create_password_ent.get()
    same_username_check=False
    if len(created_username)>5:
        messagebox.showerror('Limit Reached','Please Limit Your Username To 5 Characters Only')
    elif len(created_username)<=5:
        if create_username_ent.get() == '' and create_password_ent.get() != '':
            messagebox.showerror('Missing','Please Put Your Username')
        if create_password_ent.get() == '' and create_username_ent.get() != '':
            messagebox.showerror('Missing','Please Put Your Password')
        elif create_password_ent.get() == '' and create_username_ent.get() == '':
            messagebox.showerror('Missing','Please Check Your Username and Password Input')
        elif create_password_ent.get() != '' and create_username_ent.get() != '':
            for each_cell in excel_activate.iter_rows(values_only=True):
                if each_cell[0] == created_username:
                    same_username_check=True
            if same_username_check==True:
                messagebox.showerror('Error','Username Already Exists')
            else:
                row = excel_activate.max_row + 1
                excel_activate.cell(row=row,column=1).value = created_username
                excel_activate.cell(row=row,column=2).value = created_password
                excel_activate.cell(row=row,column=3).value = row - 1

                new_excel_activate = excel_con.create_sheet(title=created_username)
                new_excel_activate.cell(row=1,column=1).value = "Type of Dish"
                new_excel_activate.cell(row=1,column=2).value = "Name of Dish"
                new_excel_activate.cell(row=1,column=3).value = "Method of Making"
                new_excel_activate.cell(row=1,column=4).value = "Category of Dish"
                new_excel_activate.cell(row=1,column=5).value = "Ingredients"
                new_excel_activate.cell(row=1,column=6).value = "Procedure"
                new_excel_activate.cell(row=1,column=7).value = "Shopping List"
                new_excel_activate.cell(row=1,column=8).value = ""
                
                messagebox.showinfo('New Account','New Account Has Been Created Successfully')
                excel_con.save("RecipeMateDB.xlsx")
                create_username_ent.delete(0,END)
                create_password_ent.delete(0,END)
            
#===================================================================================================================================

def logout_process():
    main_interface.destroy()
    login_interface.deiconify()

#===================================================================================================================================

def main_interface_show():

    global main_interface
    login_interface.withdraw()
    main_interface = tk.Toplevel(login_interface)
    main_interface.geometry('1200x700')
    main_interface.overrideredirect(1)
    main_interface.resizable(False,False)

    global close_window_main
    def close_window_main():
        messagebox.showwarning('Exit','Please Logout First')
        
    def move_window(event):
        main_interface.geometry(f'+{event.x_root}+{event.y_root}')

    def drag(event):
        main_interface_x = main_interface.winfo_pointerx() - main_interface._offsetx
        main_interface_y = main_interface.winfo_pointery() - main_interface._offsety
        main_interface.geometry(f'+{main_interface_x}+{main_interface_y}')

    def start_drag(event):
        main_interface._offsetx = event.x
        main_interface._offsety = event.y
        
    #main interface title bar
    main_interface_title_bar=Frame(main_interface,bg='#242031')
    main_interface_title_bar.pack(fill=X)

    main_interface.bind('<Button-1>',start_drag)
    main_interface.bind('<B1-Motion>',drag)

    #main interface title bar icon
    main_interface_title_label=Label(main_interface_title_bar,text='RecipeMate (Beta) 0.0.1',fg='white',bg='#242031')
    main_interface_icon=PhotoImage(file=r'icon.png')
    main_interface_icon_pic=Label(main_interface_title_bar,image=main_interface_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
    main_interface_title_label.pack(side=LEFT,padx=10)

    #main interface close button
    main_interface_close_button=Button(main_interface_title_bar,text='✕',command=lambda:close_window_main(),bg='#242031',fg='white')
    main_interface_close_button.pack(side=RIGHT,padx=0,ipadx=7)

    window_width = 1200
    window_height = 700
    screen_width = main_interface.winfo_screenwidth()
    screen_height = main_interface.winfo_screenheight()
    x_coordinate = screen_width // 2 - window_width // 2
    y_coordinate = screen_height // 2 - window_height // 2

    main_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

    #main interface image background
    main_interface_image=Image.open(r'main_interfacebg.png')
    photo_main_interface_image=ImageTk.PhotoImage(main_interface_image)
    labelmain=Label(main_interface,image=photo_main_interface_image)
    labelmain.pack()

    global main_search_ent
    main_search_ent=Entry(main_interface,fg='#242031',bg='#F4F4F4',font=('Arial',35),highlightthickness=0,width=25,insertbackground='#242031')
    main_search_ent.place(x=250,y=292)
    main_search_ent.configure(relief='flat')

    createrecipe_photo=PhotoImage(file = r'create_recipe.png')
    createrecipe_but=Button(main_interface,image=createrecipe_photo,borderwidth=0,highlightthickness=0,command=lambda:createrecipe_interface_show())
    createrecipe_but.place(x=156,y=415)
    createrecipe_but.config(activebackground='#242031')

    showallrecipe_photo=PhotoImage(file = r'showall_recipe.png')
    showallrecipe_but=Button(main_interface,image=showallrecipe_photo,borderwidth=0,highlightthickness=0,command=lambda:showallrecipe_process())
    showallrecipe_but.place(x=666,y=415)
    showallrecipe_but.config(activebackground='#242031')
    
    def change_image():
        global index
        if index == 1:
            deleterecipe_but.config(image=deleterecipe_photo)
            searchrecipe_but.config(image=searchrecipe_icon_photo,command=lambda:search_progress())
            searchrecipe_but.place(x=884,y=277)
            index = 0
        elif index == 0:
            deleterecipe_but.config(image=searchrecipe_photo)
            searchrecipe_but.config(image=deleterecipe_icon_photo,command=lambda:delete_progress())
            searchrecipe_but.place(x=884,y=277)
            index = 1

            
    def change_iconto_edit():
        searchrecipe_but.config(image=editrecipe_photo_icon_photo,command=lambda:editrecipe_interface_show())
        searchrecipe_but.place(x=884,y=277)

    editrecipe_photo=PhotoImage(file = r'edit_recipe.png')
    editrecipe_photo_icon_photo=PhotoImage(file = r'edit_but.png')
    editrecipe_but=Button(main_interface,image=editrecipe_photo,borderwidth=0,highlightthickness=0,command=lambda:change_iconto_edit())
    editrecipe_but.place(x=156,y=545)
    editrecipe_but.config(activebackground='#242031')

    deleterecipe_photo=PhotoImage(file = r'delete_recipe.png')
    searchrecipe_icon_photo=PhotoImage(file = r'search_but.png')
    searchrecipe_photo=PhotoImage(file = r'search_recipe.png')
    deleterecipe_icon_photo=PhotoImage(file = r'delete_but.png')

    global index
    index=0

    searchrecipe_but=Button(main_interface,image=searchrecipe_icon_photo,borderwidth=0,highlightthickness=0,command=lambda:search_progress())
    searchrecipe_but.place(x=884,y=277)
    searchrecipe_but.config(activebackground='#242031')

    deleterecipe_but=Button(main_interface,image=deleterecipe_photo,borderwidth=0,highlightthickness=0,command=lambda:change_image())
    deleterecipe_but.place(x=666,y=545)
    deleterecipe_but.config(activebackground='#242031')

    def slide_animation():
        if toggle_frame.winfo_x()<0:
            for x in range(-410,1):
                toggle_frame.place(x=x)
                main_interface.update()
        else:
            for x in range(0,-410,-1):
                toggle_frame.place(x=x)
                main_interface.update()

    #Toggle Frame
    global toggle_frame
    toggle_frame=Frame(main_interface,width=400,height=674,bg='#242031')
    toggle_frame.place(x=-410,y=26)
    
    Hi_lbl=Label(toggle_frame,text='Hello',font=('Arial',35),bg='#242031',fg='white')
    Hi_lbl.place(x=155,y=14)
    
    Greeting_name=Label(toggle_frame,text=current_username,font=('Arial',35),bg='#242031',fg='white',justify='center')
    Greeting_name.place(x=155,y=84)
    
    global labeluserlogo
    userlogo_image=Image.open(r'user_logo.png')
    global userlogo_image_image2
    userlogo_image_image2=ImageTk.PhotoImage(userlogo_image)
    labeluserlogo=Label(toggle_frame,image=userlogo_image_image2)
    labeluserlogo.place(x=10,y=10)
    labeluserlogo.config(borderwidth=0,highlightthickness=0)
    
    changepic_but=Button(toggle_frame,borderwidth=5,highlightthickness=5,text="Change Picture",fg='white',bg='#242031',font=('Helvetica',20),command=lambda:change_profile_picture())
    changepic_but.place(x=5,y=160)

    social_image=Image.open(r'social.png')
    social_image_image2=ImageTk.PhotoImage(social_image)
    sociallogo=Label(toggle_frame,image=social_image_image2)
    sociallogo.place(x=50,y=250)
    sociallogo.config(borderwidth=0,highlightthickness=0)
    
    testimony_image=Image.open(r'testimony.png')
    testimony_image_image2=ImageTk.PhotoImage(testimony_image)
    testimonylogo=Label(toggle_frame,image=testimony_image_image2)
    testimonylogo.place(x=-50,y=350)
    testimonylogo.config(borderwidth=0,highlightthickness=0)
    
    logout_but=Button(toggle_frame,borderwidth=5,highlightthickness=5,text='Logout',fg='white',bg='#242031',font=('Helvetica',20),command=lambda:logout_process())
    logout_but.place(x=250,y=160)

    toggle_photo=PhotoImage(file = r'toggle_but.png')
    toggle_but=Button(main_interface,image=toggle_photo,borderwidth=0,highlightthickness=0,command=lambda:slide_animation())
    toggle_but.place(x=0,y=259)
    toggle_but.config(activebackground='#242031')

    shoppingcart_photo=PhotoImage(file = r'shopping_list.png')
    shoppingcart_but=Button(main_interface,image=shoppingcart_photo,borderwidth=0,highlightthickness=0,command=lambda:shoppinglist_interface_show())
    shoppingcart_but.place(x=1095,y=345)
    shoppingcart_but.config(activebackground='#242031')
    update_change_pic()

    main_interface.mainloop()
        
#===================================================================================================================================

def shoppinglist_interface_show():
    main_interface.withdraw()
    global shoppinglist_interface
    shoppinglist_interface = tk.Toplevel(login_interface)                        
    shoppinglist_interface.geometry('1200x700')     
    shoppinglist_interface.overrideredirect(1)
    shoppinglist_interface.resizable(False,False)

    def close_window_shoppinglist():
        shoppinglist_interface.destroy()
        main_interface.deiconify()
    def move_window(event):
        shoppinglist_interface.geometry(f'+{event.x_root}+{event.y_root}')


    def drag(event):
        shoppinglist_interface_x = shoppinglist_interface.winfo_pointerx() - shoppinglist_interface._offsetx
        shoppinglist_interface_y = shoppinglist_interface.winfo_pointery() - shoppinglist_interface._offsety
        shoppinglist_interface.geometry(f'+{shoppinglist_interface_x}+{shoppinglist_interface_y}')

    def start_drag(event):
        shoppinglist_interface._offsetx = event.x
        shoppinglist_interface._offsety = event.y
        
    # shopping list interface title bar icon
    shoppinglist_title_bar=Frame(shoppinglist_interface,bg='#242031')
    shoppinglist_title_bar.pack(fill=X)

    shoppinglist_interface.bind('<Button-1>',start_drag)
    shoppinglist_interface.bind('<B1-Motion>',drag)

    # shopping list interface title bar icon
    shoppinglist_interface_title_label=Label(shoppinglist_title_bar,text='RecipeMate (Beta) 0.0.1',fg='white',bg='#242031')
    shoppinglist_interface_icon=PhotoImage(file=r'icon.png')
    shoppinglist_interface_icon_pic=Label(shoppinglist_title_bar,image=shoppinglist_interface_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
    shoppinglist_interface_title_label.pack(side=LEFT,padx=10)

    # shopping list close button
    shoppinglist_close_button=Button(shoppinglist_title_bar,text='✕',command=lambda:close_window_shoppinglist(),bg='#242031',fg='white')
    shoppinglist_close_button.pack(side=RIGHT,padx=0,ipadx=7)

    #window configure to center
    window_width = 700
    window_height = 700
    screen_width = shoppinglist_interface.winfo_screenwidth()
    screen_height = shoppinglist_interface.winfo_screenheight()
    x_coordinate = screen_width // 2 - window_width // 2
    y_coordinate = screen_height // 2 - window_height // 2

    #window position
    shoppinglist_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

    # shopping list image background
    shoppinglist_image=Image.open(r'shoppinglist_interface.png')
    photo_shoppinglist_image=ImageTk.PhotoImage(shoppinglist_image)
    labelshoppinglist=Label(shoppinglist_interface,image=photo_shoppinglist_image)
    labelshoppinglist.pack()


    frame_shoppinglist=Frame(labelshoppinglist,width=682,height=280,bg='#242031')
    frame_shoppinglist.place(x=0,y=335)
    scrolly_shoppinglist=Scrollbar(frame_shoppinglist,orient='vertical',troughcolor='#242031',activebackground='#242031')
    scrolly_shoppinglist.pack(side=RIGHT,fill=Y)
    canvas_shoppinglist=Canvas(frame_shoppinglist,width=682,height=280,bg='#242031',yscrollcommand=scrolly_shoppinglist.set)
    canvas_shoppinglist.pack(side='left',fill='both',expand=True)
    canvas_shoppinglist.config(highlightthickness=0)
    scrolly_shoppinglist.config(command=canvas_shoppinglist.yview)
    shoppinglist_frame_content=Frame(canvas_shoppinglist,bg='#242031',relief=SOLID)
    def resize_frame(adjust):
        shoppinglist_frame_content.configure(width=adjust.width,height=adjust.height)
    shoppinglist_frame_content.bind('<Configure>',resize_frame)
    canvas_shoppinglist.create_window((0,0),window=shoppinglist_frame_content,anchor='nw')
    shoppinglist_frame_content.bind("<Configure>", lambda e: canvas_shoppinglist.configure(scrollregion=canvas_shoppinglist.bbox("all")))

    excel_con=load_workbook("RecipeMateDB.xlsx")
    account_sheet=excel_con[current_username]
    for each_cell in range(2,account_sheet.max_row + 1):
        if account_sheet['G'+str(each_cell)].value:
            nameofdishsl=account_sheet['G'+str(each_cell)].value
            ingrsl=account_sheet['H'+str(each_cell)].value
            Label(shoppinglist_frame_content, text='Dish Name: '+nameofdishsl+'\n'+'Ingredients:\n'+ingrsl,bg='#242031',fg='white',font=('Calibri',20)).pack(padx=220,pady=50)
            
    clearallbut_photo=PhotoImage(file = r'clear_shoppinglist.png')
    clearallbut=Button(labelshoppinglist,image=clearallbut_photo,borderwidth=0,highlightthickness=0,command=lambda:clear_all_sl())
    clearallbut.place(x=385,y=615)

    text_in_sl_font=font.Font(family='Arial',size=10,slant='italic')
    text_in_sl=tk.Label(labelshoppinglist,bg='#242031',fg='white',text='Tip: To Import Recipes, Try To Search The Recipe First',font=text_in_sl_font)
    text_in_sl.place(x=15,y=640)

    def clear_all_sl():
        excel_con=load_workbook('RecipeMateDB.xlsx')
        current_sheet=excel_con[current_username]
        for cell in current_sheet["G"]:
            cell.value = None
        for cell in current_sheet["H"]:
            cell.value = None
        excel_con.save("RecipeMateDB.xlsx")
        messagebox.showinfo('Success','All List Have Been Cleared')
        close_window_shoppinglist()
        shoppinglist_interface_show()

    shoppinglist_interface.mainloop()

#===================================================================================================================================

def createrecipe_interface_show():
    main_interface.withdraw()
    global createrecipe_interface
    createrecipe_interface = tk.Toplevel(main_interface)
    createrecipe_interface.geometry('1200x700')
    createrecipe_interface.overrideredirect(1)
    createrecipe_interface.resizable(False,False)

    def close_window_createrecipe():
        createrecipe_interface.destroy()
        main_interface.deiconify()
    def move_window(event):
        createrecipe_interface.geometry(f'+{event.x_root}+{event.y_root}')

    def drag(event):
        createrecipe_interface_x = createrecipe_interface.winfo_pointerx() - createrecipe_interface._offsetx
        createrecipe_interface_y = createrecipe_interface.winfo_pointery() - createrecipe_interface._offsety
        createrecipe_interface.geometry(f'+{createrecipe_interface_x}+{createrecipe_interface_y}')

    def start_drag(event):
        createrecipe_interface._offsetx = event.x
        createrecipe_interface._offsety = event.y
        
    # create recipe title bar
    createrecipe_title_bar = Frame(createrecipe_interface,bg='#242031')
    createrecipe_title_bar.pack(fill=X)

    createrecipe_interface.bind('<Button-1>',start_drag)
    createrecipe_interface.bind('<B1-Motion>',drag)

    # create recipe interface title bar icon
    createrecipe_title_label=Label(createrecipe_title_bar,text='RecipeMate (Beta) 0.0.1',fg='white',bg='#242031')
    createrecipe_icon=PhotoImage(file=r'icon.png')
    createrecipe_icon_pic=Label(createrecipe_title_bar,image=createrecipe_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
    createrecipe_title_label.pack(side=LEFT,padx=10)

    # create recipe close button
    createrecipe_close_button=Button(createrecipe_title_bar,text='✕',command=lambda:close_window_createrecipe(),bg='#242031',fg='white')
    createrecipe_close_button.pack(side=RIGHT,padx=0,ipadx=7)

    #window configure to center
    window_width = 1200
    window_height = 700
    screen_width = createrecipe_interface.winfo_screenwidth()
    screen_height = createrecipe_interface.winfo_screenheight()
    x_coordinate = screen_width // 2 - window_width // 2
    y_coordinate = screen_height // 2 - window_height // 2

    #window position
    createrecipe_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

    #create recipe interface image background
    create_interface_image=Image.open(r'createrecipe_interface.png')
    photo_create_interface_image=ImageTk.PhotoImage(create_interface_image)
    createmain=Label(createrecipe_interface,image=photo_create_interface_image)
    createmain.pack()

    #framing of contents
    frame_createrecipe=Frame(createrecipe_interface,width=1180,height=350,bg='#242031')
    frame_createrecipe.place(x=0,y=350)
    scrolly_createrecipe=Scrollbar(frame_createrecipe,orient='vertical',troughcolor='#242031',activebackground='#242031')
    scrolly_createrecipe.pack(side=RIGHT,fill=Y)
    canvas_createrecipe=Canvas(frame_createrecipe,width=1180,height=350,bg='#242031',yscrollcommand=scrolly_createrecipe.set)
    canvas_createrecipe.pack(side='left',fill='both',expand=True)
    canvas_createrecipe.config(highlightthickness=0)
    scrolly_createrecipe.config(command=canvas_createrecipe.yview)
    createrecipe_frame_content=Frame(canvas_createrecipe,bg='#242031',relief=SOLID)
    def resize_frame(adjust):
        createrecipe_frame_content.configure(width=adjust.width,height=adjust.height)
    createrecipe_frame_content.bind('<Configure>',resize_frame)
    canvas_createrecipe.create_window((0,0),window=createrecipe_frame_content,anchor='nw')

    label_spacing_vertical=15
    label_spacing_horizontal=35

    #labels
    TypeofDish_lbl=Label(createrecipe_frame_content,text='Type of Dish',font=('Arial',35),bg='#242031',fg='white')
    TypeofDish_lbl.grid(row=0,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

    global dishcombo
    global NameofDish_ent
    global mmcombo
    global cofcombo
    global ingtxt
    global proctxt


    def get(update):
        global selected_value
        NameofDish_lbl.config(text='Name of ' + str(dishcombo.get()))
        MethodofMaking_lbl.config(text='Method of Making')
        CategoryofFood_lbl.config(text='Category of ' + str(dishcombo.get()))
        Ingredients_lbl.config(text='Ingredients')
        Procedure_lbl.config(text='Procedure')
        clear_entries_recipe()

    dishvar=StringVar()
    dishlist=['Food','Drink']
    dishcombo=ttk.Combobox(createrecipe_frame_content,values=dishlist,textvariable=dishvar,font=('Helvetica',25),justify='center',state='readonly',width=35)
    dishcombo.set('Food')
    dishcombo.grid(row=0,column=3,sticky=E,columnspan=2,ipady=10)
    dishcombo.bind('<<ComboboxSelected>>',get)

    NameofDish_lbl=Label(createrecipe_frame_content,text='Name of ' + str(dishcombo.get()),font=('Arial',35),bg='#242031',fg='white')
    NameofDish_lbl.grid(row=1,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

    MethodofMaking_lbl=Label(createrecipe_frame_content,text='Method of Making',font=('Arial',35),bg='#242031',fg='white')
    MethodofMaking_lbl.grid(row=2,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

    CategoryofFood_lbl=Label(createrecipe_frame_content,text='Category of ' + str(dishcombo.get()),font=('Arial',35),bg='#242031',fg='white')
    CategoryofFood_lbl.grid(row=3,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

    Ingredients_lbl=Label(createrecipe_frame_content,text='Ingredients',font=('Arial',35),bg='#242031',fg='white')
    Ingredients_lbl.grid(row=4,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

    Procedure_lbl=Label(createrecipe_frame_content,text='Procedure',font=('Arial',35),bg='#242031',fg='white')
    Procedure_lbl.grid(row=5,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

    NameofDish_ent=Entry(createrecipe_frame_content,width=36,font=('Helvetica',25),justify=CENTER)
    NameofDish_ent.grid(row=1,column=3,sticky=E,columnspan=2,ipady=10)

    global mmlist
    mmvar=StringVar()
    mmlist=['Roasting','Baking','Grilling','Boiling','Frying','Steaming','Braising','Sauteing','Stir-frying','Microwaving','Mixing','Blending','Brewing','Fermentation','Distillation','Carbonation','Juicing']
    mmcombo=ttk.Combobox(createrecipe_frame_content,values=mmlist,textvariable=mmvar,font=('Helvetica',25),justify='center',state='readonly',width=35)
    mmcombo.current()
    mmcombo.grid(row=2,column=3,sticky=E,columnspan=2,ipady=10)

    global coflist
    cofvar=StringVar()
    coflist=['Breakfast','Snacks','Deserts','Seafood','Vegetarian/Vegan','Salads','Pasta','Non-alcoholic Beverages','Alcoholic Beverages','Hot Beverages','Cold Beverages','Carbonated Beverages','Health and Wellness Drinks','Dairy Beverages','Fruit-based Beverages','Energy Drinks']
    cofcombo=ttk.Combobox(createrecipe_frame_content,values=coflist,textvariable=cofvar,font=('Helvetica',25),justify='center',state='readonly',width=35)
    cofcombo.current()
    cofcombo.grid(row=3,column=3,sticky=E,columnspan=2,ipady=10)

    ingtxt=Text(createrecipe_frame_content,height=7,width=36,bg='white',font=('Arial',25),fg='black')
    ingtxt.grid(row=4,column=3,sticky=E,columnspan=2,ipady=10,pady=15)

    proctxt=Text(createrecipe_frame_content,height=7,width=36,bg='white',font=('Arial',25),fg='black')
    proctxt.grid(row=5,column=3,sticky=E,columnspan=2,ipady=10,pady=15)

    saverec_photo=PhotoImage(file = r'save_recipe.png')
    saverec=Button(createrecipe_frame_content,image=saverec_photo,borderwidth=0,highlightthickness=0,command=lambda:saverecipe_process())
    saverec.config(activebackground='#242031')
    saverec.grid(row=6,column=2,columnspan=2,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

    frame_createrecipe.bind('<Configure>',lambda e:canvas_createrecipe.configure(scrollregion=canvas_createrecipe.bbox('all')))

    createrecipe_interface.mainloop()
    
def saverecipe_process():
    created_dish=dishcombo.get()
    created_name=NameofDish_ent.get()
    created_method=mmcombo.get()
    created_category=cofcombo.get()
    created_ingredients=ingtxt.get('1.0','end-1c')
    created_procedure=proctxt.get('1.0','end-1c')
    
    excel_con=load_workbook("RecipeMateDB.xlsx")
    global account_sheet
    account_sheet=excel_con[current_username]  
      
    if created_name == ("") and created_method != ("") and created_category != ("") and created_ingredients != ("") and created_procedure != (""):
        messagebox.showwarning('Warning','Your Recipe Has No Name')
    if created_method == ("") and created_name != ("") and created_category != ("") and created_ingredients != ("") and created_procedure != (""):
        messagebox.showwarning('Warning','Your Recipe Has No Method')
    if created_category == ("") and created_name != ("") and created_method != ("") and created_ingredients != ("") and created_procedure != (""):
        messagebox.showwarning('Warning','Your Recipe Has No Category')
    if created_ingredients == ("")and created_name != ("") and created_method != ("") and  created_category != ("") and created_procedure != (""):
        messagebox.showwarning('Warning','Your Recipe Has No Ingredients')
    if created_procedure == ("") and created_name != ("") and created_method != ("") and  created_category != ("") and created_ingredients != (""):
        messagebox.showwarning('Warning','Your Recipe Has No Procedure')
    elif created_name == ("") and created_method == ("") and created_category == ("") and created_ingredients == ("") and created_procedure == (""):
        messagebox.showwarning('Warning','Your Recipe Has No Input, Please Fill Up All')
    elif created_name != ("") and created_method != ("") and created_category != ("") and created_ingredients != ("") and created_procedure != (""):
        progress_continue=True
        recipename_already_listed=False
        for each_cell in range(2,account_sheet.max_row + 1):
            recipename_val=account_sheet['B'+str(each_cell)].value
            if recipename_val == created_name:
                recipename_already_listed=True
                messagebox.showwarning('Already Listed','Your Recipe Is Already Listed In Your Recipe List')
                progress_continue=False
            else:
                progress_continue=True
                
        if progress_continue==True and recipename_already_listed==False:
            row_acc=account_sheet.max_row + 1
            account_sheet.cell(row=row_acc, column=1).value = created_dish
            account_sheet.cell(row=row_acc, column=2).value = created_name
            account_sheet.cell(row=row_acc, column=3).value = created_method
            account_sheet.cell(row=row_acc, column=4).value = created_category
            account_sheet.cell(row=row_acc, column=5).value = created_ingredients
            account_sheet.cell(row=row_acc, column=6).value = created_procedure
            
            excel_con.save("RecipeMateDB.xlsx")
            messagebox.showinfo('Success!', f'Your {created_name} Recipe Has Been Created')

            clear_entries_recipe()
    
#===================================================================================================================================

def editrecipe_interface_show():
    excel_con=load_workbook('RecipeMateDB.xlsx')
    current_sheet=excel_con[current_username]
    foundrecipe_edit=False
    for each_cell_edit in range(2,current_sheet.max_row + 1):
        if main_search_ent.get() == current_sheet['B'+str(each_cell_edit)].value:
            foundrecipe_edit=True
            break
    if foundrecipe_edit==True:
        main_interface.withdraw()
        global editrecipe_interface
        editrecipe_interface = tk.Toplevel(main_interface)
        editrecipe_interface.geometry('1200x700')
        editrecipe_interface.overrideredirect(1)
        editrecipe_interface.resizable(False,False)

        def close_window_editrecipe():
            editrecipe_interface.destroy()
            main_interface.deiconify()

        def move_window(event):
            editrecipe_interface.geometry(f'+{event.x_root}+{event.y_root}')

        def drag(event):
            editrecipe_interface_x = editrecipe_interface.winfo_pointerx() - editrecipe_interface._offsetx
            editrecipe_interface_y = editrecipe_interface.winfo_pointery() - editrecipe_interface._offsety
            editrecipe_interface.geometry(f'+{editrecipe_interface_x}+{editrecipe_interface_y}')

        def start_drag(event):
            editrecipe_interface._offsetx = event.x
            editrecipe_interface._offsety = event.y
            
        # edit recipe title bar
        editrecipe_title_bar = Frame(editrecipe_interface,bg='#242031')
        editrecipe_title_bar.pack(fill=X)

        editrecipe_interface.bind('<Button-1>',start_drag)
        editrecipe_interface.bind('<B1-Motion>',drag)

        # edit recipe interface title bar icon
        editrecipe_title_label=Label(editrecipe_title_bar,text='RecipeMate (Beta) 0.0.1',fg='white',bg='#242031')
        editrecipe_icon=PhotoImage(file=r'icon.png')
        editrecipe_icon_pic=Label(editrecipe_title_bar,image=editrecipe_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
        editrecipe_title_label.pack(side=LEFT,padx=10)

        # edit recipe close button
        editrecipe_close_button=Button(editrecipe_title_bar,text='✕',command=lambda:close_window_editrecipe(),bg='#242031',fg='white')
        editrecipe_close_button.pack(side=RIGHT,padx=0,ipadx=7)

        #window configure to center
        window_width = 1200
        window_height = 700
        screen_width = editrecipe_interface.winfo_screenwidth()
        screen_height = editrecipe_interface.winfo_screenheight()
        x_coordinate = screen_width // 2 - window_width // 2
        y_coordinate = screen_height // 2 - window_height // 2

        #window position
        editrecipe_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

        #edit recipe interface image background
        create_interface_image=Image.open(r'editrecipe_interface.png')
        photo_create_interface_image=ImageTk.PhotoImage(create_interface_image)
        editbg=Label(editrecipe_interface,image=photo_create_interface_image)
        editbg.pack()

        #framing of contents
        frame_editrecipe=Frame(editrecipe_interface,width=1180,height=350,bg='#242031')
        frame_editrecipe.place(x=0,y=350)
        scrolly_editrecipe=Scrollbar(frame_editrecipe,orient='vertical',troughcolor='#242031',activebackground='#242031')
        scrolly_editrecipe.pack(side=RIGHT,fill=Y)
        canvas_editrecipe=Canvas(frame_editrecipe,width=1180,height=350,bg='#242031',yscrollcommand=scrolly_editrecipe.set)
        canvas_editrecipe.pack(side='left',fill='both',expand=True)
        canvas_editrecipe.config(highlightthickness=0)
        scrolly_editrecipe.config(command=canvas_editrecipe.yview)
        frame_content=Frame(canvas_editrecipe,bg='#242031',relief=SOLID)
        def resize_frame(adjust):
            frame_content.configure(width=adjust.width,height=adjust.height)
        frame_content.bind('<Configure>',resize_frame)
        canvas_editrecipe.create_window((0,0),window=frame_content,anchor='nw')

        global label_spacing_vertical
        global label_spacing_horizontal
        label_spacing_vertical=15
        label_spacing_horizontal=35

        #labels
        TypeofDish_lbl=Label(frame_content,text='Type of Dish',font=('Arial',35),bg='#242031',fg='white')
        TypeofDish_lbl.grid(row=0,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)



        def get(update):
            global selected_value
            NameofDish_lbl.config(text='Name of ' + str(edit_dishvar.get()))
            MethodofMaking_lbl.config(text='Method of Making')
            CategoryofFood_lbl.config(text='Category of ' + str(edit_dishvar.get()))
            Ingredients_lbl.config(text='Ingredients')
            Procedure_lbl.config(text='Procedure')


        newtypedish=StringVar()
        newnamedish=StringVar()
        newmm=StringVar()
        newcof=StringVar()


        edit_dishvar=['Food','Drink']
        edit_dishvar=ttk.Combobox(frame_content,values=edit_dishvar,textvariable=newtypedish,font=('Helvetica',25),justify='center',state='readonly',width=35)
        edit_dishvar.set('Food')
        edit_dishvar.grid(row=0,column=3,sticky=E,columnspan=2,ipady=10)
        edit_dishvar.bind('<<ComboboxSelected>>',get)
        #===============================
        dishint=IntVar()
        dishChk=Checkbutton(frame_content,text="same as before",variable=dishint,bg='#242031',font=15,fg='white',command=lambda:get_existing_typeofdish())
        dishChk.grid(row=1,column=3,sticky=E)
        #==============================
        NameofDish_lbl=Label(frame_content,text='Name of ' + str(edit_dishvar.get()),font=('Arial',35),bg='#242031',fg='white')
        NameofDish_lbl.grid(row=2,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

        MethodofMaking_lbl=Label(frame_content,text='Method of Making',font=('Arial',35),bg='#242031',fg='white')
        MethodofMaking_lbl.grid(row=4,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

        CategoryofFood_lbl=Label(frame_content,text='Category of ' + str(edit_dishvar.get()),font=('Arial',35),bg='#242031',fg='white')
        CategoryofFood_lbl.grid(row=6,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

        Ingredients_lbl=Label(frame_content,text='Ingredients',font=('Arial',35),bg='#242031',fg='white')
        Ingredients_lbl.grid(row=8,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

        Procedure_lbl=Label(frame_content,text='Procedure',font=('Arial',35),bg='#242031',fg='white')
        Procedure_lbl.grid(row=10,column=0,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

        edit_NameofDish_ent=Entry(frame_content,width=36,font=('Helvetica',25),justify=CENTER,textvariable=newnamedish)
        edit_NameofDish_ent.grid(row=2,column=3,sticky=E,columnspan=2,ipady=10)
        #==============================
        namedishint=IntVar()
        namedishChk=Checkbutton(frame_content,text="same as before",variable=namedishint,bg='#242031',font=15,fg='white',command=lambda:get_existing_nameoffood())
        namedishChk.grid(row=3,column=3,sticky=E)  
        #==============================
        edit_mmlist=['Roasting','Baking','Grilling','Boiling','Frying','Steaming','Braising','Sauteing','Stir-frying','Microwaving','Mixing','Blending','Brewing','Fermentation','Distillation','Carbonation','Juicing']
        edit_mmcombo=ttk.Combobox(frame_content,values=edit_mmlist,textvariable=newmm,font=('Helvetica',25),justify='center',state='readonly',width=35)
        edit_mmcombo.current()
        edit_mmcombo.grid(row=4,column=3,sticky=E,columnspan=2,ipady=10)
        #==============================
        mmint=IntVar()
        mmChk=Checkbutton(frame_content,text="same as before",variable=mmint,bg='#242031',fg='white',font=15,command=lambda:get_existing_methodofmaking())
        mmChk.grid(row=5,column=3,sticky=E)  
        #==============================
        edit_coflist=['Breakfast','Snacks','Deserts','Seafood','Vegetarian/Vegan','Salads','Pasta','Non-alcoholic Beverages','Alcoholic Beverages','Hot Beverages','Cold Beverages','Carbonated Beverages','Health and Wellness Drinks','Dairy Beverages','Fruit-based Beverages','Energy Drinks']
        edit_cofcombo=ttk.Combobox(frame_content,values=edit_coflist,textvariable=newcof,font=('Helvetica',25),justify='center',state='readonly',width=35)
        edit_cofcombo.current()
        edit_cofcombo.grid(row=6,column=3,sticky=E,columnspan=2,ipady=10)
        #==============================
        cofint=IntVar()
        cofChk=Checkbutton(frame_content,text="same as before",variable=cofint,bg='#242031',fg='white',font=15,command=lambda:get_existing_categoryoffood())
        cofChk.grid(row=7,column=3,sticky=E) 
        #==============================
        edit_ingtxt=Text(frame_content,height=7,width=36,bg='white',font=('Arial',25),fg='black')
        edit_ingtxt.grid(row=8,column=3,sticky=E,columnspan=2,ipady=10,pady=15)
        #==============================
        ingint=IntVar()
        ingChk=Checkbutton(frame_content,text="same as before",variable=ingint,bg='#242031',fg='white',font=15,command=lambda:get_existing_ingredients())
        ingChk.grid(row=9,column=3,sticky=E) 
        #==============================
        edit_proctxt=Text(frame_content,height=7,width=36,bg='white',font=('Arial',25),fg='black')
        edit_proctxt.grid(row=10,column=3,sticky=E,columnspan=2,ipady=10,pady=15)
        #==============================
        procint=IntVar()
        procChk=Checkbutton(frame_content,text="same as before",variable=procint,bg='#242031',fg='white',font=15,command=lambda:get_existing_procedure())
        procChk.grid(row=11,column=3,sticky=E) 
        #==============================
        editrec_photo=PhotoImage(file = r'edit_recipe2.png')
        editrec=Button(frame_content,image=editrec_photo,borderwidth=0,highlightthickness=0,command=lambda:update_prog())
        editrec.config(activebackground='#242031')
        editrec.grid(row=12,column=3,columnspan=2,pady=label_spacing_vertical,padx=label_spacing_horizontal,sticky=W)

        account_sheet=excel_con[current_username]

        def get_existing_typeofdish():
            if dishint.get() == 1:
                olddishtype = account_sheet['A' + str(each_cell_edit)].value
                newtypedish.set(olddishtype)
            elif dishint.get() == 0:
                newtypedish.set('')
        def get_existing_nameoffood():
            if namedishint.get() == 1:
                global oldnamedish
                oldnamedish = account_sheet['B' + str(each_cell_edit)].value
                newnamedish.set(oldnamedish)
            elif namedishint.get() == 0:
                newnamedish.set('')
        def get_existing_methodofmaking():
            if mmint.get() == 1:
                oldmm = account_sheet['C' + str(each_cell_edit)].value
                newmm.set(oldmm)
            elif mmint.get() == 0:
                newmm.set('')
        def get_existing_categoryoffood():
            if cofint.get() == 1:
                oldcof = account_sheet['D' + str(each_cell_edit)].value
                newcof.set(oldcof)
            elif cofint.get() == 0:
                newcof.set('')
        def get_existing_ingredients():
            if ingint.get() == 1:
                oldingr = account_sheet['E' + str(each_cell_edit)].value
                edit_ingtxt.insert(END,str(oldingr))
            elif ingint.get() == 0:
                edit_ingtxt.delete(1.0,END)

        def get_existing_procedure():
            if procint.get() == 1:
                oldproc = account_sheet['F' + str(each_cell_edit)].value
                edit_proctxt.insert(END,str(oldproc))
            elif procint.get() == 0:
                edit_proctxt.delete(1.0,END)
            
        def clear_entries_editrecipe():
            edit_NameofDish_ent.delete(0,END)
            edit_mmcombo.set("")
            edit_cofcombo.set("")
            edit_ingtxt.delete(1.0,END)
            edit_proctxt.delete(1.0,END)
        def update_prog():
            created_name=edit_NameofDish_ent.get()
            created_method=edit_mmcombo.get()
            created_category=edit_cofcombo.get()
            created_ingredients=edit_ingtxt.get('1.0','end-1c')
            created_procedure=edit_proctxt.get('1.0','end-1c')
            if created_name == ("") and created_method != ("") and created_category != ("") and created_ingredients != ("") and created_procedure != (""):
                messagebox.showwarning('Warning','Your Recipe Has No Name')
            if created_method == ("") and created_name != ("") and created_category != ("") and created_ingredients != ("") and created_procedure != (""):
                messagebox.showwarning('Warning','Your Recipe Has No Method')
            if created_category == ("") and created_name != ("") and created_method != ("") and created_ingredients != ("") and created_procedure != (""):
                messagebox.showwarning('Warning','Your Recipe Has No Category')
            if created_ingredients == ("")and created_name != ("") and created_method != ("") and  created_category != ("") and created_procedure != (""):
                messagebox.showwarning('Warning','Your Recipe Has No Ingredients')
            if created_procedure == ("") and created_name != ("") and created_method != ("") and  created_category != ("") and created_ingredients != (""):
                messagebox.showwarning('Warning','Your Recipe Has No Procedure')
            elif created_name == ("") and created_method == ("") and created_category == ("") and created_ingredients == ("") and created_procedure == (""):
                messagebox.showwarning('Warning','Your Recipe Has No Input, Please Fill Up All')
            elif created_name != ("") and created_method != ("") and created_category != ("") and created_ingredients != ("") and created_procedure != (""):
                progress_continue=True
                recipename_already_listed=False               
                for each_cell in range(2,account_sheet.max_row + 1):
                    recipename_val=account_sheet['B'+str(each_cell)].value
                    if recipename_val == created_name and oldnamedish != created_name:
                        recipename_already_listed=True
                        messagebox.showwarning('Already Listed','Your Recipe Is Already Listed In Your Recipe List')
                        progress_continue=False            
                    else:
                        progress_continue=True
                
                if progress_continue==True and recipename_already_listed==False or oldnamedish == created_name:
                    account_sheet['A' + str(each_cell_edit)].value = edit_dishvar.get()
                    account_sheet['B' + str(each_cell_edit)].value = edit_NameofDish_ent.get()
                    account_sheet['C' + str(each_cell_edit)].value = edit_mmcombo.get()
                    account_sheet['D' + str(each_cell_edit)].value = edit_cofcombo.get()
                    account_sheet['E' + str(each_cell_edit)].value = edit_ingtxt.get('1.0','end-1c')
                    account_sheet['F' + str(each_cell_edit)].value = edit_proctxt.get('1.0','end-1c')
                    messagebox.showinfo('Success!', f'Your {created_name} Recipe Has Been Updated')
                    excel_con.save("RecipeMateDB.xlsx")  
                    clear_entries_editrecipe()
                    close_window_editrecipe()          
    else:
        messagebox.showerror('Recipe Not Found','Searched Recipe Is Not Found')

    frame_editrecipe.bind('<Configure>',lambda e:canvas_editrecipe.configure(scrollregion=canvas_editrecipe.bbox('all')))

    editrecipe_interface.mainloop()

#===================================================================================================================================

def search_progress():
    foundrecipe_search=False
    excel_con=load_workbook('RecipeMateDB.xlsx')
    current_sheet=excel_con[current_username]
    for each_cell in range(2,current_sheet.max_row + 1):
        if main_search_ent.get() == current_sheet['B'+str(each_cell)].value:
            recipe_recordnum=int(each_cell) - 1
            foundrecipe_search=True
            break
    clear_entries_main()
    if foundrecipe_search==True:
        main_interface.withdraw()
        messagebox.showinfo('Recipe Found',"Recipe Found!! Loading Interface...")
        searchresult_interface = tk.Toplevel(main_interface)
        searchresult_interface.geometry('1200x700')
        searchresult_interface.overrideredirect(1)
        searchresult_interface.resizable(False,False)

        def close_window_searchresult():
            searchresult_interface.destroy()
            main_interface.deiconify()

        def move_window(event):
            searchresult_interface.geometry(f'+{event.x_root}+{event.y_root}')

        def drag(event):
            searchresult_interface_x = searchresult_interface.winfo_pointerx() - searchresult_interface._offsetx
            searchresult_interface_y = searchresult_interface.winfo_pointery() - searchresult_interface._offsety
            searchresult_interface.geometry(f'+{searchresult_interface_x}+{searchresult_interface_y}')

        def start_drag(event):
            searchresult_interface._offsetx = event.x
            searchresult_interface._offsety = event.y
            
        # search result title bar
        searchresult_title_bar = Frame(searchresult_interface,bg='#242031')
        searchresult_title_bar.pack(fill=X)

        searchresult_interface.bind('<Button-1>',start_drag)
        searchresult_interface.bind('<B1-Motion>',drag)

        # search result interface title bar icon
        searchresult_title_label=Label(searchresult_title_bar,text='RecipeMate (Beta) 0.0.1',fg='white',bg='#242031')
        searchresult_icon=PhotoImage(file=r'icon.png')
        searchresult_icon_pic=Label(searchresult_title_bar,image=searchresult_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
        searchresult_title_label.pack(side=LEFT,padx=10)

        # search result close button
        searchresult_close_button=Button(searchresult_title_bar,text='✕',command=lambda:close_window_searchresult(),bg='#242031',fg='white')
        searchresult_close_button.pack(side=RIGHT,padx=0,ipadx=7)

        #window configure to center
        window_width = 1200
        window_height = 700
        screen_width = searchresult_interface.winfo_screenwidth()
        screen_height = searchresult_interface.winfo_screenheight()
        x_coordinate = screen_width // 2 - window_width // 2
        y_coordinate = screen_height // 2 - window_height // 2

        #window position
        searchresult_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

        #search result interface image background
        create_interface_image=Image.open(r'searchrecipe_interface2.png')
        photo_create_interface_image=ImageTk.PhotoImage(create_interface_image)
        searchresbg=Label(searchresult_interface,image=photo_create_interface_image)
        searchresbg.pack()
        
        frame_searchres=Frame(searchresbg,width=650,height=617,bg='#242031')
        frame_searchres.place(x=525,y=20)
        scrolly_searchres=Scrollbar(frame_searchres,orient='vertical',troughcolor='#242031',activebackground='#242031')
        scrolly_searchres.pack(side=RIGHT,fill=Y)
        canvas_searchres=Canvas(frame_searchres,width=650,height=617,bg='#242031',yscrollcommand=scrolly_searchres.set)
        canvas_searchres.pack(side='left',fill='both',expand=True)
        canvas_searchres.config(highlightthickness=0)
        scrolly_searchres.config(command=canvas_searchres.yview)
        searchres_frame_content=Frame(canvas_searchres,bg='#242031',relief=SOLID)
        def resize_frame(adjust):
            searchres_frame_content.configure(width=adjust.width,height=adjust.height)
        searchres_frame_content.bind('<Configure>',resize_frame)
        canvas_searchres.create_window((0,0),window=searchres_frame_content,anchor='nw')
        searchres_frame_content.bind("<Configure>", lambda e: canvas_searchres.configure(scrollregion=canvas_searchres.bbox("all")))
        
        
        txtishtype = current_sheet['A' + str(each_cell)].value
        txtnamedish = current_sheet['B' + str(each_cell)].value
        txtmm = current_sheet['C' + str(each_cell)].value
        txtcof = current_sheet['D' + str(each_cell)].value
        txtingr = current_sheet['E' + str(each_cell)].value
        txtproc = current_sheet['F' + str(each_cell)].value

        TypeofDish_lbl=Label(searchres_frame_content,text='Type of Dish',font=('Arial',35),bg='#242031',fg='#FFDE59')
        TypeofDish_lbl.pack(pady=15,padx=175)

        TypeofDish_value=Label(searchres_frame_content,text=txtishtype,font=('Arial',20),bg='#242031',fg='white')
        TypeofDish_value.pack(pady=15)
        
        NameofDish_lbl=Label(searchres_frame_content,text='Name of Dish',font=('Arial',35),bg='#242031',fg='#FFDE59')
        NameofDish_lbl.pack(pady=15)

        NameofDish_value=Label(searchres_frame_content,text=txtnamedish,font=('Arial',20),bg='#242031',fg='white')
        NameofDish_value.pack(pady=15)

        MethodofMaking_lbl=Label(searchres_frame_content,text='Method of Making',font=('Arial',35),bg='#242031',fg='#FFDE59')
        MethodofMaking_lbl.pack(pady=15)

        MethodofMaking_value=Label(searchres_frame_content,text=txtmm,font=('Arial',20),bg='#242031',fg='white')
        MethodofMaking_value.pack(pady=15)

        CategoryofFood_lbl=Label(searchres_frame_content,text='Category of Dish',font=('Arial',35),bg='#242031',fg='#FFDE59')
        CategoryofFood_lbl.pack(pady=15)

        CategoryofFood_value=Label(searchres_frame_content,text=txtcof,font=('Arial',20),bg='#242031',fg='white')
        CategoryofFood_value.pack(pady=15)
        
        Ingredients_lbl=Label(searchres_frame_content,text='Ingredients',font=('Arial',35),bg='#242031',fg='#FFDE59')
        Ingredients_lbl.pack(pady=15)

        Ingredients_value=Label(searchres_frame_content,text=txtingr,font=('Arial',20),bg='#242031',fg='white',justify=LEFT)
        Ingredients_value.pack(pady=15)
        
        Procedure_lbl=Label(searchres_frame_content,text='Procedure',font=('Arial',35),bg='#242031',fg='#FFDE59')
        Procedure_lbl.pack(pady=15)
        
        Procedure_value=Label(searchres_frame_content,text=txtproc,font=('Arial',20),bg='#242031',fg='white',justify=LEFT)
        Procedure_value.pack(pady=15)        
        
        addtocart_photo=PhotoImage(file = r'add_to_cart.png')
        Addtocart_but=Button(searchres_frame_content,image=addtocart_photo,borderwidth=0,highlightthickness=0,command=lambda:add_to_shopping_list())
        Addtocart_but.pack(pady=33)
        
        
        def add_to_shopping_list():
            txtnamedish
            txtingr
            excel_con=load_workbook("RecipeMateDB.xlsx")
            account_sheet=excel_con[current_username]
            already_listed=False
            for each_cell in range(2,account_sheet.max_row + 1):
                Gvalue = account_sheet['G'+str(each_cell)].value
                if Gvalue == txtnamedish:
                    already_listed=True
                    messagebox.showwarning('Already Listed','Your Recipe Is Already Listed In Your Shopping List')
                    progress_continue=False
                else:
                    progress_continue=True
                    
            if progress_continue==True and already_listed==False:
                row_acc=account_sheet.max_row + 1
                account_sheet.cell(row=row_acc, column=7).value = txtnamedish
                account_sheet.cell(row=row_acc, column=8).value = txtingr
                excel_con.save("RecipeMateDB.xlsx")
                messagebox.showinfo('Success!', f'Your {txtnamedish} Recipe Has Been Imported To Your Own Shopping List')

        searchresult_interface.mainloop()
    else:
        messagebox.showerror('Recipe Not Found','Searched Recipe is not found')
    clear_entries_main()
    
#===================================================================================================================================

def delete_progress():
    foundrecipe_delete=False
    excel_con=load_workbook('RecipeMateDB.xlsx')
    current_sheet=excel_con[current_username]
    for each_cell in range(2,current_sheet.max_row + 1):
        if main_search_ent.get() == current_sheet['B'+str(each_cell)].value:
            cell_address = each_cell
            foundrecipe_delete=True
            break
    if foundrecipe_delete==True:
        current_sheet.delete_rows(cell_address)
        messagebox.showinfo("Recipe Deleted","Recipe Has Been Deleted Successfully")
        excel_con.save("RecipeMateDB.xlsx")
    else:
        messagebox.showerror("Recipe Not Deleted ","Recipe Does Not Exists")
    clear_entries_main()
     
#===================================================================================================================================

def showallrecipe_process():
    main_interface.withdraw()
    global showallrecipe_interface
    showallrecipe_interface = tk.Toplevel(main_interface)
    showallrecipe_interface.geometry('1200x700')
    showallrecipe_interface.overrideredirect(1)
    showallrecipe_interface.resizable(False,False)

    def close_window_showallrecipe():
        showallrecipe_interface.destroy()
        main_interface.deiconify()

    def move_window(event):
        showallrecipe_interface.geometry(f'+{event.x_root}+{event.y_root}')

    def drag(event):
        showallrecipe_interface_x = showallrecipe_interface.winfo_pointerx() - showallrecipe_interface._offsetx
        showallrecipe_interface_y = showallrecipe_interface.winfo_pointery() - showallrecipe_interface._offsety
        showallrecipe_interface.geometry(f'+{showallrecipe_interface_x}+{showallrecipe_interface_y}')

    def start_drag(event):
        showallrecipe_interface._offsetx = event.x
        showallrecipe_interface._offsety = event.y
        
    # show all recipe title bar
    showallrecipe_title_bar = Frame(showallrecipe_interface,bg='#242031')
    showallrecipe_title_bar.pack(fill=X)

    showallrecipe_interface.bind('<Button-1>',start_drag)
    showallrecipe_interface.bind('<B1-Motion>',drag)

    # show all recipe interface title bar icon
    showallrecipe_title_label=Label(showallrecipe_title_bar,text='RecipeMate (Beta) 0.0.1',fg='white',bg='#242031')
    showallrecipe_icon=PhotoImage(file=r'icon.png')
    showallrecipe_icon_pic=Label(showallrecipe_title_bar,image=showallrecipe_icon,highlightthickness=0,borderwidth=0,bg='#242031').pack(side=LEFT)
    showallrecipe_title_label.pack(side=LEFT,padx=10)

    # show all recipe close button
    showallrecipe_close_button=Button(showallrecipe_title_bar,text='✕',command=lambda:close_window_showallrecipe(),bg='#242031',fg='white')
    showallrecipe_close_button.pack(side=RIGHT,padx=0,ipadx=7)

    #window configure to center
    window_width = 1200
    window_height = 700
    screen_width = showallrecipe_interface.winfo_screenwidth()
    screen_height = showallrecipe_interface.winfo_screenheight()
    x_coordinate = screen_width // 2 - window_width // 2
    y_coordinate = screen_height // 2 - window_height // 2

    #window position
    showallrecipe_interface.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')

    #show all recipe interface image background
    create_interface_image=Image.open(r'searchrecipe_interface.png')
    photo_create_interface_image=ImageTk.PhotoImage(create_interface_image)
    showallbg=Label(showallrecipe_interface,image=photo_create_interface_image)
    showallbg.pack()

    #framing of contents
    frame_showallrecipe=Frame(showallrecipe_interface,width=800,height=10,bg='#242031')
    frame_showallrecipe.place(x=170,y=50)
    
    frame_showallrecipe_withbut=Frame(showallrecipe_interface,width=400,height=50,bg='#242031')
    frame_showallrecipe_withbut.place(x=515,y=645)
    
    Refresh_but=Button(frame_showallrecipe_withbut,borderwidth=5,highlightthickness=5,text='Refresh',fg='white',bg='#242031',font=('Helvetica',15),command=lambda:refresh_showallrecipe_process())
    Refresh_but.grid(row=0,column=0)
    
    global combobox
    combobox_values = ["All Recipes","Food","Drink"]
    combobox = ttk.Combobox(frame_showallrecipe_withbut, values=combobox_values)
    combobox.bind("<<ComboboxSelected>>", categorize_treeview)
    combobox.grid(row=0,column=1)
    combobox.set("All Recipes")
    
    global tv1
    global treescrolly
    global treescrollx
    tv1 = ttk.Treeview(frame_showallrecipe,height=3)
    treescrolly = ttk.Scrollbar(frame_showallrecipe, orient="vertical", command=tv1.yview)
    treescrollx = ttk.Scrollbar(frame_showallrecipe, orient="horizontal", command=tv1.xview)
    tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
    treescrollx.pack(side ="bottom",fill ="x")
    treescrolly.pack(side ="right",fill="y")  

    Record_no=0
    tv1['columns'] = ("Type of Dish","Name of Dish","Method of Making","Category of Dish","Ingredients","Procedure")
    tv1.column("#0", width=0)
    tv1.column("Type of Dish", anchor=CENTER, width=90, stretch=True)
    tv1.column("Name of Dish",  anchor=CENTER, width=90, stretch=True)
    tv1.column("Method of Making", anchor=CENTER, width=120, stretch=True)
    tv1.column("Category of Dish", anchor=CENTER, width=120, stretch=True)
    tv1.column("Ingredients", anchor=W, width=200, stretch=True)
    tv1.column("Procedure", anchor=W, width=200, stretch=True)
    
    tv1.heading("#0", text="", anchor=CENTER)
    tv1.heading("Type of Dish", text="Type of Dish", anchor=CENTER)
    tv1.heading("Name of Dish", text="Name of Dish", anchor=CENTER)
    tv1.heading("Method of Making", text="Method of Making", anchor=CENTER)
    tv1.heading("Category of Dish", text="Category of Dish", anchor=CENTER)
    tv1.heading("Ingredients", text="Ingredients", anchor=CENTER)
    tv1.heading("Procedure", text="Procedure", anchor=CENTER)
    
    style = ttk.Style(main_interface)
    style.configure('Treeview', rowheight=177, background='white', foreground='#242031')
    

    excel_con=load_workbook('RecipeMateDB.xlsx')
    current_sheet=excel_con[current_username]
    for each_cell in range(2, (current_sheet.max_row)+1):
        Record_no+=1
        tv1.insert(parent='', index="end", values=(current_sheet['A'+str(each_cell)].value,current_sheet['B'+str(each_cell)].value,current_sheet['C'+str(each_cell)].value,current_sheet['D'+str(each_cell)].value,current_sheet['E'+str(each_cell)].value,current_sheet['F'+str(each_cell)].value))
    tv1.pack()
    
    
    showallrecipe_interface.mainloop()

def refresh_showallrecipe_process():
    tv1.destroy()
    treescrolly.destroy()
    treescrollx.destroy()
    showallrecipe_process()
    
def categorize_treeview(update):
    
    excel_con=load_workbook('RecipeMateDB.xlsx')
    current_sheet=excel_con[current_username]
    selected_value = combobox.get()
    if selected_value == 'All Recipes':
        tv1.delete(*tv1.get_children())
        for row in current_sheet.iter_rows(min_row=2, values_only=True):
            if 'Drink' in row[0] or 'Food' in row[0]:
                tv1.insert('', 'end', values=row)
    if selected_value == 'Drink':
        tv1.delete(*tv1.get_children())
        for row in current_sheet.iter_rows(min_row=2, values_only=True):
            if 'Drink' in row[0] and row[0] is not None:
                tv1.insert('', 'end', values=row)
    if selected_value == 'Food':
        tv1.delete(*tv1.get_children())
        for row in current_sheet.iter_rows(min_row=2, values_only=True):
            if 'Food' in row[0] and row[0] is not None:
                tv1.insert('', 'end', values=row)
                
#===================================================================================================================================

def change_profile_picture():
    
    profilepic_path = filedialog.askopenfilename(filetypes=[("image files", "*.png;*.jpg;*.jpeg;*.gif")])
    if profilepic_path:
        global new_photo
        img=Image.open(profilepic_path)
        img=img.resize((135, 135))
        new_photo=ImageTk.PhotoImage(img)

        excel_file="RecipeMateDB.xlsx"
        excel_con=load_workbook(excel_file)
        account_sheet = excel_con[current_username]
        global new_img
        
        #remove old
        if len(account_sheet._images) > 1:
            account_sheet._images.clear()
            
        new_img=xlImage(profilepic_path)
        account_sheet.add_image(new_img, 'L1')
        
        new_img=account_sheet['L1']
        new_img.value=None
        
        excel_con.save(excel_file)
        
        messagebox.showinfo('Success!', 'You Have Updated Your Profile Picture. Please Log in Again.')
        
        main_interface.destroy()
        login_interface.deiconify()
        update_change_pic()
    else:
        messagebox.showerror('Error!', 'No Image File Selected.')

def update_change_pic():
    excel_con=load_workbook('RecipeMateDB.xlsx')
    sheet_current=excel_con[current_username]

    image_loader=SheetImageLoader(sheet_current)
    image_path=image_loader.get('L1')

    photo_image=image_path.resize((135, 135))
    imagel1=ImageTk.PhotoImage(photo_image)

    labeluserlogo.config(image=imagel1)
    labeluserlogo.image=imagel1

#===================================================================================================================================

#welcome interface image background
loading_interface_image=Image.open(r'Loading_interface_bg.png')
photo_loading_interface_image=ImageTk.PhotoImage(loading_interface_image)
welcome_label=Label(loading_interface,image=photo_loading_interface_image)
welcome_label.pack()
welcome_text_font=font.Font(family='Arial',size=15,slant='italic')
text_in_loading=tk.Label(loading_interface,bg='#242031',fg='white',text='',font=welcome_text_font)
text_in_loading.place(x=142,y=210)

def pick_number_loading():
    number = random.randint(1,5)
    if number == 1:
        text = 'Preparing Dishes'
        text_in_loading.place(x=95,y=210)
    elif number == 2:
        text = 'Planning all Menu'
        text_in_loading.place(x=91,y=210)
    elif number == 3:
        text = 'Someone Seems Hungry'
        text_in_loading.place(x=60,y=210)
    elif number == 4:
        text = 'Welcome User!'
        text_in_loading.place(x=104,y=210)
    else:
        text = 'Guessing Your Next Meal'
        text_in_loading.place(x=60,y=210)
    text_in_loading.config(text=text)

pick_number_loading()
fade_out()

tk.mainloop()